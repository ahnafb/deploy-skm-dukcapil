from flask import (Flask, json, render_template, jsonify, request, redirect, url_for, send_file, session)
from pymongo import MongoClient
from unittest import result
import jwt
import hashlib
import csv
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from bson import ObjectId
import pandas as pd
from flask_cors import CORS 
from dotenv import load_dotenv
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

app = Flask(__name__)
CORS(app)

# Load .env file
load_dotenv()

# Read environment variables
connection_string = os.getenv("DB_CONNECTION_STRING")
db_name = os.getenv("DB_NAME")
SECRET_KEY = os.getenv("SECRET_KEY")

# Set up MongoDB connection
client = MongoClient(connection_string)
db = client[db_name]

@app.route('/')
def home():
    # Ambil data jenis layanan dari database
    services = db.services.find()
    genders = db.genders.find()
    ages = db.ages.find() 
    pendidikans = db.pendidikans.find()     
    pekerjaans = db.pekerjaans.find()
    questions = db.questions.find()
    pertanyaan = db.pertanyaan.find()
    # Kirim data ke template HTML
    return render_template('index.html', pertanyaan=pertanyaan, services=services, genders=genders, ages=ages, pendidikans=pendidikans, pekerjaans=pekerjaans, questions=questions)

@app.route('/datadiri', methods=['GET'])
def show_datadiri():
    sample_receive = request.args.get('sample_give')
    print(sample_receive)
    return jsonify({'msg': 'GET request complete!'})

@app.route('/datadiri', methods=['POST'])
def save_datadiri():
    layanan_receive = request.form.get("layanan_give")
    kelamin_receive = request.form.get("kelamin_give")
    usia_receive = request.form.get("usia_give")
    pendidikan_receive = request.form.get("pendidikan_give")
    pekerjaan_receive = request.form.get("pekerjaan_give")
    if pekerjaan_receive == "Lainnya":
        pekerjaan_receive = request.form.get("pekerjaan_lainnya")

    jawaban_pertanyaan = request.form.get('jawaban_pertanyaan')
    if jawaban_pertanyaan:
        jawaban_pertanyaan = json.loads(jawaban_pertanyaan)
    
    # Calculate total score
    total_skor = sum([jawaban['nilai'] for jawaban in jawaban_pertanyaan.values()])

    kritik_receive = request.form.get("kritik")
    saran_receive = request.form.get("saran")
    tanggal_dibuat = datetime.now().strftime("%Y-%m-%d || %H:%M:%S")

    doc = {
        'layanan': layanan_receive,
        'jenis_kelamin': kelamin_receive,
        'usia': usia_receive,
        'pendidikan': pendidikan_receive,
        'pekerjaan': pekerjaan_receive,
        'jawaban_pertanyaan': jawaban_pertanyaan,
        'total_skor': total_skor,  # Add total score to document
        'boxKritik': kritik_receive,
        'boxSaran': saran_receive,
        'waktu_dibuat': tanggal_dibuat,
    }
    db.TData30.insert_one(doc)
    return jsonify({'msg': 'POST request complete!', 'total_skor': total_skor})  # Include total score in response

@app.route('/admin-page', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        jenis_layanan_selected = request.form.get('jenis_layanan')
        if jenis_layanan_selected == 'all':
            datadiris = db.TData30.find()  # Memuat semua data jika dipilih "Semua Jenis Layanan"
        else:
            datadiris = db.TData30.find({'layanan': jenis_layanan_selected})  # Memuat data sesuai dengan jenis layanan yang dipilih
        return render_template('table_rows.html', datadiris=datadiris,)  # Mengirim kembali data tabel yang diperbarui
    else:
        datadiris = db.TData30.find()  # Mengambil semua data dari koleksi datadiri
        services = db.services.find()  # Memuat semua data layanan untuk ditampilkan di halaman admin
        genders = db.genders.find()  # Memuat semua data jenis kelamin untuk ditampilkan di halaman admin
        return render_template('admin-page.html', datadiris=datadiris, services=services, genders=genders)

@app.route('/statistik_dashboard')
def statistik_dashboard():
    # Mengambil data dari database
    data = list(db.TData30.find({}, {'_id': 0}))  # Mengambil semua data kecuali _id

    # Menghitung jumlah jenis kelamin
    jenis_kelamin_count = {}
    layanan_count = {}  # Untuk menghitung jumlah tiap jenis layanan

    for item in data:
        # Menghitung jenis kelamin
        jenis_kelamin = item.get('jenis_kelamin')
        if jenis_kelamin:
            if jenis_kelamin in jenis_kelamin_count:
                jenis_kelamin_count[jenis_kelamin] += 1
            else:
                jenis_kelamin_count[jenis_kelamin] = 1

        # Menghitung jenis layanan
        layanan = item.get('layanan')
        if layanan:
            if layanan in layanan_count:
                layanan_count[layanan] += 1
            else:
                layanan_count[layanan] = 1

    return render_template('statistik_dashboard.html', jenis_kelamin_count=jenis_kelamin_count, layanan_count=layanan_count)

# !!!!!!! Form Page Tampil CRUD
@app.route('/form_dashboard')
def form():
    services = db.services.find()  
    genders = db.genders.find()
    ages = db.ages.find() 
    pendidikans = db.pendidikans.find()
    pekerjaans = db.pekerjaans.find()
    questions = db.questions.find()
    pertanyaan = db.pertanyaan.find()
    return render_template('form_dashboard.html', services=services, genders=genders, ages=ages, pendidikans=pendidikans, pekerjaans=pekerjaans, questions=questions, pertanyaan=pertanyaan)

# !!!!!!! Form page ->> CRUD Jenis Layanan
@app.route('/add_service', methods=['POST'])
def add_service():
    service_name = request.form['service_name']
    db.services.insert_one({'name': service_name})
    return jsonify({'message': 'Jenis Layanan berhasil ditambahkan!'})

@app.route('/edit_service', methods=['POST'])
def edit_service():
    service_id = request.form['service_id']
    new_name = request.form['new_name']
    db.services.update_one({'_id': ObjectId(service_id)}, {'$set': {'name': new_name}})
    return jsonify({'message': 'Jenis Layanan berhasil di-update!'})

@app.route('/delete_service', methods=['POST'])
def delete_service():
    service_id = request.form['service_id']
    db.services.delete_one({'_id': ObjectId(service_id)})
    return jsonify({'message': 'Jenis Layanan berhasil dihapus!'})  

# !!!!!!! Form page ->> CRUD Jenis Kelamin
@app.route('/add_gender', methods=['POST'])
def add_gender():
    gender_name = request.form['gender_name']
    db.genders.insert_one({'name': gender_name})
    return jsonify({'message': 'Jenis Kelamin berhasil ditambahkan!'})

@app.route('/edit_gender', methods=['POST'])
def edit_gender():
    gender_id = request.form['gender_id']
    new_name = request.form['new_name']
    db.genders.update_one({'_id': ObjectId(gender_id)}, {'$set': {'name': new_name}})
    return jsonify({'message': 'Jenis Kelamin berhasil di-update!'})

@app.route('/delete_gender', methods=['POST'])
def delete_gender():
    gender_id = request.form['gender_id']
    db.genders.delete_one({'_id': ObjectId(gender_id)})
    return jsonify({'message': 'Jenis Kelamin berhasil dihapus!'})

# !!!!!!! Form page ->> CRUD Usia
# Route untuk menambah usia
@app.route('/add_age', methods=['POST'])
def add_age():
    age_range = request.form['age_range']
    db.ages.insert_one({'range': age_range})
    return jsonify({'message': 'Usia berhasil ditambahkan!'})

# Route untuk mengedit usia
@app.route('/edit_age', methods=['POST'])
def edit_age():
    age_id = request.form['age_id']
    new_range = request.form['new_range']
    db.ages.update_one({'_id': ObjectId(age_id)}, {'$set': {'range': new_range}})
    return jsonify({'message': 'Usia berhasil di-update!'})

# Route untuk menghapus usia
@app.route('/delete_age', methods=['POST'])
def delete_age():
    age_id = request.form['age_id']
    db.ages.delete_one({'_id': ObjectId(age_id)})
    return jsonify({'message': 'Usia berhasil dihapus!'})

# !!!!!!! Form page ->> CRUD Pendidikan
@app.route('/add_pendidikan', methods=['POST'])
def add_pendidikan():
    pendidikan_name = request.form['pendidikan_name']
    db.pendidikans.insert_one({'name': pendidikan_name})
    return jsonify({'message': 'Pendidikan berhasil ditambahkan!'})

# Route untuk mengedit pendidikan
@app.route('/edit_pendidikan', methods=['POST'])
def edit_pendidikan():
    pendidikan_id = request.form['pendidikan_id']
    new_name = request.form['new_name']
    db.pendidikans.update_one({'_id': ObjectId(pendidikan_id)}, {'$set': {'name': new_name}})
    return jsonify({'message': 'Pendidikan berhasil di-update!'})

# Route untuk menghapus pendidikan
@app.route('/delete_pendidikan', methods=['POST'])
def delete_pendidikan():
    pendidikan_id = request.form['pendidikan_id']
    db.pendidikans.delete_one({'_id': ObjectId(pendidikan_id)})
    return jsonify({'message': 'Pendidikan berhasil dihapus!'})

# !!!!!!! Form page ->> CRUD Pendidikan
# Route untuk menambah pekerjaan
@app.route('/add_pekerjaan', methods=['POST'])
def add_pekerjaan():
    pekerjaan_name = request.form['pekerjaan_name']
    db.pekerjaans.insert_one({'name': pekerjaan_name})
    return jsonify({'message': 'Pekerjaan berhasil ditambahkan!'})

# Route untuk mengedit pekerjaan
@app.route('/edit_pekerjaan', methods=['POST'])
def edit_pekerjaan():
    pekerjaan_id = request.form['pekerjaan_id']
    new_name = request.form['new_name']
    db.pekerjaans.update_one({'_id': ObjectId(pekerjaan_id)}, {'$set': {'name': new_name}})
    return jsonify({'message': 'Pekerjaan berhasil di-update!'})

# Route untuk menghapus pekerjaan
@app.route('/delete_pekerjaan', methods=['POST'])
def delete_pekerjaan():
    pekerjaan_id = request.form['pekerjaan_id']
    db.pekerjaans.delete_one({'_id': ObjectId(pekerjaan_id)})
    return jsonify({'message': 'Pekerjaan berhasil dihapus!'})

# #################################################33
@app.route('/add_question', methods=['POST'])
def add_question():
    # Ambil data dari permintaan POST
    pertanyaan = request.form['pertanyaan']
    nilai = float(request.form['nilai'])  # Ubah ke float
    
    pilihan1 = request.form['pilihan1']    # Ambil nilai dari pilihan 1
    pilihan2 = request.form['pilihan2']    # Ambil nilai dari pilihan 2
    pilihan3 = request.form['pilihan3']    # Ambil nilai dari pilihan 3
    pilihan4 = request.form['pilihan4']    # Ambil nilai dari pilihan 4
    
    nilai_pilihan1 = request.form['nilai_pilihan1']  # Ambil nilai dari pilihan 1
    nilai_pilihan2 = request.form['nilai_pilihan2']  # Ambil nilai dari pilihan 2
    nilai_pilihan3 = request.form['nilai_pilihan3']  # Ambil nilai dari pilihan 3
    nilai_pilihan4 = request.form['nilai_pilihan4']  # Ambil nilai dari pilihan 4
    
    # Tambahkan pertanyaan ke database
    db.pertanyaan.insert_one({
        'pertanyaan': pertanyaan,
        'nilai': nilai,
        'pilihan1': pilihan1,
        'pilihan2': pilihan2,
        'pilihan3': pilihan3,
        'pilihan4': pilihan4,
        'nilai_pilihan1': float(nilai_pilihan1),  # Ubah ke float jika perlu
        'nilai_pilihan2': float(nilai_pilihan2),  # Ubah ke float jika perlu
        'nilai_pilihan3': float(nilai_pilihan3),  # Ubah ke float jika perlu
        'nilai_pilihan4': float(nilai_pilihan4),  # Ubah ke float jika perlu
    })
    
    # Kirim respons JSON ke klien
    return jsonify({'message': 'Pertanyaan berhasil ditambahkan!'})

@app.route('/edit_question', methods=['POST'])
def edit_question():
    # Ambil data dari permintaan POST
    question_id = request.form['question_id']
    edited_question = request.form['edited_question']
    edited_option1 = request.form['edited_option1']
    edited_option2 = request.form['edited_option2']
    edited_option3 = request.form['edited_option3']
    edited_option4 = request.form['edited_option4']

    # Perbarui pertanyaan di database
    db.pertanyaan.update_one(
        {'_id': ObjectId(question_id)},
        {'$set': {
            'pertanyaan': edited_question,
            'pilihan1': edited_option1,
            'pilihan2': edited_option2,
            'pilihan3': edited_option3,
            'pilihan4': edited_option4
        }}
    )

    # Kirim respons JSON ke klien
    return jsonify({'message': 'Pertanyaan berhasil diubah.'})

@app.route('/delete_question', methods=['POST'])
def delete_question():
    # Ambil data dari permintaan POST
    question_id = request.form['question_id']

    # Hapus pertanyaan dari database
    db.pertanyaan.delete_one({'_id': ObjectId(question_id)})

    # Kirim respons JSON ke klien
    return jsonify({'message': 'Pertanyaan berhasil dihapus!'})


# !!!!!! end of admin page

# ! Version1 Excel
@app.route('/export-excel', methods=['GET'])
def export_excel():
    # Ambil jenis layanan dan rentang waktu dari parameter URL
    jenis_layanan = request.args.get('jenis_layanan', 'all')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Konversi tanggal dari string ke objek datetime
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)

    # Query untuk mengambil data dari MongoDB sesuai jenis layanan dan rentang waktu
    query = {}
    if jenis_layanan != 'all':
        query['layanan'] = jenis_layanan

    if start_date and end_date:
        query['waktu_dibuat'] = {'$gte': start_date, '$lte': end_date}

    data = list(db.TData30.find(query, {'_id': 0}))

    # Buat dataframe dari data MongoDB
    df = pd.DataFrame(data)

    # Menyesuaikan header
    custom_header = {
        'layanan': 'Jenis Layanan',
        'jenis_kelamin': 'Jenis Kelamin',
        'usia': 'Usia',
        'pendidikan': 'Pendidikan',
        'pekerjaan': 'Pekerjaan',
        'pekerjaanLainnya': 'Pekerjaan Lainnya',
        'boxKritik': 'Kritik',
        'boxSaran': 'Saran',
        'total_skor': 'Total Skor',
        'jawaban_pertanyaan': 'Jawaban Pertanyaan'
    }
    df = df.rename(columns=custom_header)

    # Konversi kolom yang berisi list atau dictionary menjadi string
    def convert_to_strings(value):
        if isinstance(value, (list, dict)):
            return json.dumps(value)
        return value

    df = df.applymap(convert_to_strings)

    # Nama file Excel sesuai dengan jenis layanan dan rentang waktu
    if jenis_layanan == 'all':
        filename = 'all-data.xlsx'
    else:
        filename = jenis_layanan.replace(' ', '-') + '-data.xlsx'

    # Buat file Excel dengan openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Tambahkan judul
    title = "Data Hasil Survey Kepuasan Masyarakat Dispendukcapil Kota Semarang"
    if start_date and end_date:
        subtitle = f"Jenis Layanan: {jenis_layanan.replace('-', ' ')} | Periode: {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}"
    else:
        subtitle = f"Jenis Layanan: {jenis_layanan.replace('-', ' ')} | Periode: Semua"

    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)

    ws.merge_cells('A2:J2')
    ws['A2'] = subtitle
    ws['A2'].font = Font(size=12, bold=True)

    # Tambahkan data dari DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Simpan file Excel
    wb.save(filename)

    # Kirim file Excel ke pengguna
    return send_file(filename, as_attachment=True)

# ! Start of Login
@app.route("/login", methods=['GET'])
def login():
    msg = request.args.get("msg")
    return render_template('login.html', msg=msg)

@app.route('/sign_up/check_dup', methods=['POST'])
def check_dup():
    username_receive = request.form['username_give']
    exists = bool(db.users.find_one({"username": username_receive}))
    return jsonify({'result': 'success', 'exists': exists})

@app.route("/sign_up/save", methods=["POST"])
def sign_up():
    username_receive = request.form['username_give']
    password_receive = request.form['password_give']
    password_hash = hashlib.sha256(password_receive.encode('utf-8')).hexdigest()
    doc = {
        "username": username_receive,                               # id
        "password": password_hash,                                  # password
        "profile_name": username_receive,                           # user's name is set to their id by default
        "profile_pic": "",                                          # profile image file name
        "profile_pic_real": "profile_pics/profile_placeholder.png", # a default profile image
        "profile_info": ""                                          # a profile description
    }
    db.users.insert_one(doc)
    return jsonify({'result': 'success'})

@app.route("/sign_in", methods=["POST"])
def sign_in():
    # Sign in
    username_receive = request.form["username_give"]
    password_receive = request.form["password_give"]
    pw_hash = hashlib.sha256(password_receive.encode("utf-8")).hexdigest()
    result = db.users.find_one(
        {
            "username": username_receive,
            "password": pw_hash,
        }
    )
    if result:
        payload = {
            "id": username_receive,
            # the token will be valid for 24 hours
            "exp": datetime.utcnow() + timedelta(seconds=60 * 60 * 24),
        }
        token = jwt.encode(payload, SECRET_KEY, algorithm="HS256")

        return jsonify(
            {
                "result": "success",
                "token": token,
            }
        )
    # Let's also handle the case where the id and
    # password combination cannot be found
    else:
        return jsonify(
            {
                "result": "fail",
                "msg": "We could not find a user with that id/password combination",
            }
        )
        
@app.route('/sign_out')
def sign_out():
    if 'username' in session:
        session.pop('username')
    return redirect(url_for('home'))

# ! End of Login

# ! Endpoint Periode Waktu
@app.route('/periode', methods=['GET', 'POST'])
def periode():
    datadiris = []
    if request.method == 'POST':
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        
        # Convert the dates from string to datetime objects
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)

        # Query the database for entries within the date range
        all_data = db.TData30.find()
        
        for data in all_data:
            if 'waktu_dibuat' in data:
                waktu_dibuat = datetime.strptime(data['waktu_dibuat'], '%Y-%m-%d || %H:%M:%S')
                if start_date <= waktu_dibuat <= end_date:
                    datadiris.append(data)

    return render_template('admin-page.html', datadiris=datadiris)

if __name__ == '__main__':
    app.run('0.0.0.0', port=5000, debug=True)